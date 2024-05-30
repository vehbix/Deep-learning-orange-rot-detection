import tensorflow as tf
import os
import cv2
import numpy as np
from PIL import Image
import random
from matplotlib import pyplot as plt
from tensorflow.keras.models import load_model
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay

# Modeli yükleme
model = load_model('models/imageclassifier3.h5')

# Rastgele görüntü seçimi için dizinleri belirleyin
curuk_dir = 'data/curuk0'
duzgun_dir = 'data/duzgun0'

# Her iki dizinden de rastgele görüntüler seçin
curuk_images = random.sample(os.listdir(curuk_dir), 20)
duzgun_images = random.sample(os.listdir(duzgun_dir), 20)

# Tam dosya yollarını oluşturun
liste = [os.path.join(curuk_dir, img) for img in curuk_images] + \
        [os.path.join(duzgun_dir, img) for img in duzgun_images]

def detect_rotten_areas(image):
    # Görüntüyü LAB renk alanına dönüştürme
    lab = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)

    # Portakal rengi için renk eşikleme
    lower_orange = np.array([20, 120, 120])
    upper_orange = np.array([40, 200, 255])
    mask_orange = cv2.inRange(lab, lower_orange, upper_orange)
    
    # Portakal rengi maskesini çürük bölgelerin maskeleriyle birleştirme
    combined_mask = mask_orange
    
    # Morfolojik işlemlerle maskeyi iyileştirme
    kernel = np.ones((5, 5), np.uint8)
    combined_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
    combined_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_OPEN, kernel)

    # Çürük bölgeleri tespit etmek için renk eşikleme
    lower_rotten1 = np.array([10, 10, 10])
    upper_rotten1 = np.array([100, 135, 135])
    
    lower_rotten2 = np.array([135, 120, 110])
    upper_rotten2 = np.array([200, 150, 145])
    
    lower_rotten3 = np.array([135, 120, 110])
    upper_rotten3 = np.array([255, 160, 150])
    
    lower_rotten4 = np.array([55, 70, 70])
    upper_rotten4 = np.array([135, 160, 150])

    mask_rotten1 = cv2.inRange(lab, lower_rotten1, upper_rotten1)
    mask_rotten2 = cv2.inRange(lab, lower_rotten2, upper_rotten2)
    mask_rotten3 = cv2.inRange(lab, lower_rotten3, upper_rotten3)
    mask_rotten4 = cv2.inRange(lab, lower_rotten4, upper_rotten4)

    # Maskeleri birleştirme
    combined_rotten_mask = cv2.bitwise_or(mask_rotten1, mask_rotten2)
    combined_rotten_mask = cv2.bitwise_or(combined_rotten_mask, mask_rotten3)
    combined_rotten_mask = cv2.bitwise_or(combined_rotten_mask, mask_rotten4)

    # Maske üzerinde morfolojik işlemler uygulama
    combined_rotten_mask = cv2.morphologyEx(combined_rotten_mask, cv2.MORPH_CLOSE, kernel)
    combined_rotten_mask = cv2.morphologyEx(combined_rotten_mask, cv2.MORPH_OPEN, kernel)

    # Konturları bulma
    contours, _ = cv2.findContours(combined_rotten_mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    # Konturları çizme
    output_image = image.copy()
    cv2.drawContours(output_image, contours, -1, (0, 255, 0), 2)

    return output_image, combined_rotten_mask, contours

def calculate_area_percentage(contours, image_path):
    image = Image.open(image_path)
    image_rgb = image.convert("RGB")
    image_array_rgb = np.array(image_rgb)

    # Siyah ve yeşil renklerin RGB değerleri
    black_pixel_rgb = [0, 0, 0]
    green_pixel_rgb = [0, 255, 0]

    # Siyah ve yeşil olmayan piksellerin sayısını hesaplama
    non_black_green_mask = np.all(image_array_rgb != black_pixel_rgb, axis=-1) & np.all(image_array_rgb != green_pixel_rgb, axis=-1)
    total_area = np.sum(non_black_green_mask)

    # Çürük alanı hesaplama
    rotten_area = sum(cv2.contourArea(contour) for contour in contours)
    # Yüzdelik alanı hesapla
    percentage = (rotten_area / total_area) * 101.12

    if percentage > 98:
        percentage = 100
    return percentage

# Excel dosyasını oluştur ve çalışma sayfasını seç
wb = Workbook()
ws = wb.active
ws.title = "Rotten Area Detection"

# Başlıkları ekle
headers = ["Original Image", "Detected Rotten Areas", "Percentage of Rotten Area", "Prediction"]
for col_num, header in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    ws[f"{col_letter}1"] = header
    ws[f"{col_letter}1"].border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

# Hücre boyutlarını ayarla
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
image_height = 220  # Görüntülerin yüksekliği (piksel)
row_height = image_height // 1  # Excel satır yüksekliğini ayarlama (ortalama oran)

for i, image_path in enumerate(liste):
    print(image_path)
    img = cv2.imread(image_path)
    resize = cv2.resize(img, (256, 256))

    resim = resize.copy()
    yhat = model.predict(np.expand_dims(resim / 255, 0))
    print(yhat)

    row = i + 2
    if yhat < 0.13:
        print(f'Predicted class is Çürük')
        # Çürük bölgeleri tespit et ve göster
        detected_image, mask, contours = detect_rotten_areas(img)
        percentage = calculate_area_percentage(contours, image_path)
        print(f'Rotten area: {percentage:.2f}%')

        # Orijinal ve tespit edilen görüntüleri kaydet
        original_image_path = f'crk/original_image_{i}.png'
        detected_image_path = f'dtc/detected_image_{i}.png'
        cv2.imwrite(original_image_path, img)
        cv2.imwrite(detected_image_path, detected_image)

        # Orijinal görüntüyü ekle
        img_original = OpenpyxlImage(original_image_path)
        img_original.anchor = f"A{row}"
        ws.add_image(img_original)

        # Tespit edilen görüntüyü ekle
        img_detected = OpenpyxlImage(detected_image_path)
        img_detected.anchor = f"B{row}"
        ws.add_image(img_detected)

        # Yüzdelik alanı ve tahmin sonucunu ekle
        ws[f"C{row}"] = f"{percentage:.2f}%"
        ws[f"D{row}"] = "Çürük"
        
    else:
        print(f'Predicted class is Düzgün')

        # Orijinal görüntüyü kaydet
        original_image_path = f'duzg/original_image_{i}.png'
        cv2.imwrite(original_image_path, img)

        # Orijinal görüntüyü ekle
        img_original = OpenpyxlImage(original_image_path)
        img_original.anchor = f"A{row}"
        ws.add_image(img_original)

        # Tespit edilen görüntü olarak aynı orijinal görüntüyü ekle
        img_detected = OpenpyxlImage(original_image_path)
        img_detected.anchor = f"B{row}"
        ws.add_image(img_detected)

        # Yüzdelik alanı ve tahmin sonucunu ekle
        ws[f"C{row}"] = "0.00%"
        ws[f"D{row}"] = "Düzgün"

    # Hücre kenarlıkları ekle
    for col_num in range(1, 5):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}{row}"].border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        ws[f"{col_letter}{row}"].alignment = Alignment(horizontal='center', vertical='center')
    
    # Satır yüksekliğini ayarla
    ws.row_dimensions[row].height = row_height

# Workbook'u kaydet
wb.save('resim_with_rotten_areas.xlsx')


# Veriyi yükleme ve normalleştirme
data = tf.keras.utils.image_dataset_from_directory('data', batch_size=1, shuffle=True)
data = data.map(lambda x, y: (x / 255, y))

# Veriyi eğitim, doğrulama ve test setlerine ayırma
train_size = int(len(data) * 0.7)
val_size = int(len(data) * 0.2)
test_size = int(len(data) * 0.1)

train = data.take(train_size)
val = data.skip(train_size).take(val_size)
test = data.skip(train_size + val_size).take(test_size)

# Eğitim, doğrulama ve test verileri üzerindeki gerçek ve tahmin edilen etiketleri toplama
y_pred_dl = []
y_true_dl = []

for images, labels in data:  # Tüm veri seti üzerinde döngü
    y_pred_batch = model.predict(images)
    y_pred_dl.extend(np.round(y_pred_batch).astype(int).flatten())
    y_true_dl.extend(labels.numpy().astype(int))

# Karışıklık matrisini çizme
cm = confusion_matrix(y_true_dl, y_pred_dl)
disp = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=['Çürük','Düzgün'])
disp.plot(cmap=plt.cm.Blues)
plt.title('Confusion Matrix for All Data')
plt.show()
